from __future__ import annotations

from collections import Counter
from datetime import datetime
from io import BytesIO

import pytest
from openpyxl import load_workbook
from openpyxl import Workbook as OpenpyxlWorkbook

from irradiation_analysis import generator as generator_module
from irradiation_analysis.analytics import find_growth_signals
from irradiation_analysis.excel_io import (
    OPTIONAL_COLUMNS,
    REQUIRED_COLUMNS,
    UploadedWorkbook,
    import_workbooks,
)
from irradiation_analysis.generator import (
    DEFAULT_MONITOR_TYPES,
    MonitorTypeConfig,
    SimulationConfig,
    build_blank_template,
    build_prefilled_template,
    generate_simulated_workbooks,
)
from irradiation_analysis.snapshots import all_device_ids
from irradiation_analysis.models import MonitoringStatus
from irradiation_analysis.status import classify_record


DATA_SHEET = "监测数据"


def workbook_from_bytes(content: bytes):
    return load_workbook(BytesIO(content), data_only=True)


def test_blank_template_has_required_headers_and_frozen_pane():
    workbook = workbook_from_bytes(build_blank_template())
    worksheet = workbook[DATA_SHEET]

    headers = [cell.value for cell in worksheet[1]]

    assert headers == list(REQUIRED_COLUMNS + OPTIONAL_COLUMNS)
    assert worksheet.freeze_panes == "A2"
    assert worksheet.auto_filter.ref == "A1:L2"
    assert worksheet.column_dimensions["A"].width >= 18
    assert worksheet.column_dimensions["E"].width >= 12
    assert worksheet["A2"].number_format == "yyyy-mm-dd hh:mm"
    assert worksheet["E2"].number_format == "0.000"
    assert "设备阈值配置" in workbook.sheetnames

    workbook.close()


def test_prefilled_template_contains_two_hundred_devices_per_monitor_type():
    monitor_types = (
        MonitorTypeConfig("γ剂量率", "μSv/h", 10.0, 20.0),
        MonitorTypeConfig("中子剂量率", "μSv/h", 5.0, 12.0),
    )
    config = SimulationConfig(
        start=datetime(2026, 6, 1, 8),
        end=datetime(2026, 6, 1, 8),
        monitor_types=monitor_types,
        seed=123,
    )

    content = build_prefilled_template(config)
    workbook = workbook_from_bytes(content)
    worksheet = workbook[DATA_SHEET]
    rows = list(
        worksheet.iter_rows(
            min_row=2,
            values_only=True,
        )
    )
    monitor_type_counts = Counter(row[3] for row in rows if row[0] is not None)
    device_ids = {row[2] for row in rows if row[0] is not None}

    assert len(rows) == 400
    assert monitor_type_counts == {"γ剂量率": 200, "中子剂量率": 200}
    assert device_ids == set(all_device_ids())

    result = import_workbooks([UploadedWorkbook("prefilled.xlsx", content)])
    assert result.summary.valid_rows == 400
    assert result.summary.blocked_rows == 0

    workbook.close()


def test_simulation_is_reproducible_and_importable():
    config = SimulationConfig(
        start=datetime(2026, 6, 1),
        end=datetime(2026, 6, 6),
        sampling_hours=24,
        warning_ratio=0.05,
        accident_ratio=0.03,
        rapid_growth_ratio=0.04,
        event_duration=3,
        seed=20260616,
    )

    first = generate_simulated_workbooks(config)
    second = generate_simulated_workbooks(config)

    assert [(workbook.filename, workbook.content) for workbook in first] == [
        (workbook.filename, workbook.content) for workbook in second
    ]

    result = import_workbooks(first)

    assert len(first) == 1
    assert first[0].filename == "simulated_monitoring_20260601_20260606.xlsx"
    assert result.summary.blocked_rows == 0
    assert result.summary.valid_rows > 0
    assert set(DEFAULT_MONITOR_TYPES).issubset(result.summary.monitor_types)


def test_simulation_includes_requested_statuses_and_consistent_thresholds():
    config = SimulationConfig(
        start=datetime(2026, 6, 1),
        end=datetime(2026, 6, 6),
        sampling_hours=24,
        warning_ratio=0.05,
        accident_ratio=0.03,
        rapid_growth_ratio=0.04,
        event_duration=3,
        seed=42,
    )
    result = import_workbooks(generate_simulated_workbooks(config))
    statuses = {classify_record(record) for record in result.records}

    assert statuses == {
        MonitoringStatus.NORMAL,
        MonitoringStatus.WARNING,
        MonitoringStatus.ACCIDENT,
    }
    assert all(
        0 < record.warning_threshold < record.control_threshold
        for record in result.records
    )
    assert find_growth_signals(result.records)


def test_long_range_rapid_growth_simulation_triggers_growth_signals():
    config = SimulationConfig(
        start=datetime(2026, 1, 1),
        end=datetime(2026, 3, 5),
        sampling_hours=24,
        monitor_types=(MonitorTypeConfig("dose-rate", "uSv/h", 10.0, 20.0),),
        warning_ratio=0,
        accident_ratio=0,
        rapid_growth_ratio=1.0,
        event_duration=3,
        seed=20260616,
    )

    result = import_workbooks(generate_simulated_workbooks(config))
    signals = find_growth_signals(result.records)

    assert result.summary.blocked_rows == 0
    assert result.summary.valid_rows == 64 * len(all_device_ids())
    assert signals
    assert all(signal.latest_value > signal.previous_value for signal in signals)


def test_overallocated_scenario_ratios_raise_clear_error():
    config = SimulationConfig(
        start=datetime(2026, 6, 1),
        end=datetime(2026, 6, 1),
        warning_ratio=0.8,
        accident_ratio=0.8,
        rapid_growth_ratio=0.8,
        seed=123,
    )

    with pytest.raises(ValueError, match="scenario ratios"):
        generate_simulated_workbooks(config)


@pytest.mark.filterwarnings("error")
def test_monthly_simulation_uses_streaming_workbook_and_remains_importable(
    monkeypatch,
):
    workbook_calls: list[dict[str, object]] = []

    def workbook_spy(*args, **kwargs):
        workbook_calls.append(dict(kwargs))
        return OpenpyxlWorkbook(*args, **kwargs)

    monkeypatch.setattr(generator_module, "Workbook", workbook_spy)
    config = SimulationConfig(
        start=datetime(2026, 6, 1),
        end=datetime(2026, 7, 1),
        sampling_hours=24,
        warning_ratio=0.05,
        accident_ratio=0.03,
        rapid_growth_ratio=0.04,
        event_duration=3,
        seed=20260616,
    )

    workbooks = generate_simulated_workbooks(config)
    result = import_workbooks(workbooks)

    assert workbook_calls[0].get("write_only") is True
    assert len(workbooks) == 1
    assert result.summary.blocked_rows == 0
    assert result.summary.valid_rows == 31 * len(all_device_ids()) * len(DEFAULT_MONITOR_TYPES)
    assert find_growth_signals(result.records)


def test_unknown_output_mode_raises_clear_error():
    config = SimulationConfig(
        start=datetime(2026, 6, 1),
        end=datetime(2026, 6, 1),
        output_mode="weekly",
    )

    with pytest.raises(ValueError, match="output_mode"):
        generate_simulated_workbooks(config)


@pytest.mark.parametrize(
    ("field", "value", "message"),
    [
        ("start", "2026-06-01", "start must be a datetime"),
        ("end", "2026-06-01", "end must be a datetime"),
        ("sampling_hours", 1.5, "sampling_hours must be a positive integer"),
        ("event_duration", 2.5, "event_duration must be a positive integer"),
        ("warning_ratio", "0.5", "warning_ratio must be a finite number"),
    ],
)
def test_invalid_config_types_raise_clear_errors(field, value, message):
    kwargs = {
        "start": datetime(2026, 6, 1),
        "end": datetime(2026, 6, 1),
    }
    kwargs[field] = value
    config = SimulationConfig(**kwargs)

    with pytest.raises(ValueError, match=message):
        generate_simulated_workbooks(config)
